from board import Board
from tuple import Tuple
import numpy as np
import copy
import os
import openpyxl

if __name__ == "__main__":

    #creat Excel
    exl = openpyxl.Workbook()
    exl.save('Tuple1.xlsx')
    exl = openpyxl.load_workbook('Tuple1.xlsx') 
    sheet1 = exl['Sheet']
    sheet1.cell(1, 1).value = 'Game number' #(row, col) = (1, 1)
    sheet1.cell(1, 2).value = 'Win or lose'
    sheet1.cell(1, 3).value = 'Final Score'
    exl.save('Tuple1.xlsx')

    
    ntuple = Tuple()
    
    winCot = 0
    loseCot = 0
    #highestScore = 0
    #playCot = 0 # record how many board had played
    #n = 0
    list = ["U", "R", "D", "L"]

    k = 200
    final = np.zeros((k, 3), dtype = np.uint32)
    
    for j in range(k):
        game = Board()
        winFlag = 0
        score = 0
        cin_flag = -1
        #final[i][0] = j
        while True:
            #score, cin_flag, n = game.checkWinLoseContinue(game, ntuple, cin_flag, score, n)
            tmp = Board()
            tmax = -1
            irecord = -1
            scorelist = [0, 0, 0, 0]

            for i in range(4):
                tmp.board = game.board
                m = tmp.move(list[i])
                scorelist[i] = m
                checksum = tmp.check()
                if checksum == -5:
                    continue
                elif checksum == 100:
                    winFlag = 1
                    irecord = i
                    break
                else:
                    t = ntuple.getTupleValueSum(tmp.board)
                    if t > tmax:
                        tmax = t
                        irecord = i
            
            if winFlag == 1:
                print("win")
                score += game.move(list[i])
                cin_flag = 1 #win or lose
                break
            elif tmax == -1: #四個方向都不能走 -> 遊戲結束
                print("lose")
                maxscore = -1
                irecordm = -1
                for w in range(4):
                    if scorelist[w] > maxscore:
                        maxscore = scorelist[w]
                        irecordm = w
                score += game.move(list[irecordm])
                cin_flag = 0
                break
            else:
                score += game.move(list[irecord])
                game.newTile()
                game.printBoard()
                print("Score: ", score)
            

        if cin_flag == 0:
            loseCot += 1
        elif cin_flag == 1:
            winCot += 1

        final[j][0] = j
        final[j][1] = cin_flag
        final[j][2] = score
        
        sheet1.cell(j+2, 1).value = j
        sheet1.cell(j+2, 2).value = cin_flag
        sheet1.cell(j+2, 3).value = score
        exl.save('Tuple1.xlsx')
    
    print(winCot / k)
    print(final)
    sheet1.cell(k+5, 1).value = (winCot / k)
    exl.save('Tuple1.xlsx')
        #exl.save


