from board import Board
from tuple import Tuple
import numpy as np
import copy
import os
import openpyxl
import csv
import time

if __name__ == "__main__":

    #creat Excel -> record results（win or lose）
    exl = openpyxl.Workbook()
    exl.save('Tuple2.xlsx')
    exl = openpyxl.load_workbook('Tuple2.xlsx') 
    sheet1 = exl['Sheet']
    sheet1.cell(1, 1).value = 'Game number' #(row, col) = (1, 1)
    sheet1.cell(1, 2).value = 'Win or lose'
    sheet1.cell(1, 3).value = 'Final Score'
    exl.save('Tuple2.xlsx')

    # record results（% of tiles）
    filename = open('output.txt', 'w',newline='')
    
    ntuple = Tuple()
    game = Board()
    
    winCot = 0
    loseCot = 0

    list = ["R", "D", "L", "U"]

    k = 1000 #episode
    milestone = 100
    final = np.zeros((k, 3), dtype = np.uint32)
    score200 = []
    score200sum = 0

    totalReward = 0
    
    for j in range(k):
        original_tuple = np.copy(ntuple.tuple)
        game.startNewGame()
        score = 0
        cin_flag = 0
        r, game.board = ntuple.takeAction(list, game)

        while True:
            getT1 = ntuple.getTupleValueSum(game.board)
            #r, game.board = ntuple.takeAction(list, game)
            totalReward += r
            score += r

            # getT2 = ntuple.getTupleValueSum(game.board)
            # ntuple.udpTuple(game.board, r + (getT2 - getT1))

            # if game.check() == 100:
            #     cin_flag = 1
            game.newTile()
            r, game.board = ntuple.takeAction(list, game)

            if game.check() == 100:
                cin_flag = 1
            if r == -1:
                break
            
            # if r == -1:
            #     cin_flag = 0
            #     if game.check() == 100:
            #         cin_flag = 1
            #     break
            
            # if r != -1:
            #     game.newTile()
            
            # if game.check() == -5:
            #     cin_flag = 0
            #     break

            getT2 = ntuple.getTupleValueSum(game.board)
            ntuple.udpTuple(game.board, r + (getT2 - getT1))

        # record result (% of tiles)
        ntuple.addCount(game)  

        game.printBoard()
        print("Score: ", score)
        print("k: ", j)
        
        if cin_flag == 0:
            loseCot += 1
        elif cin_flag == 1:
            winCot += 1

        final[j][0] = j
        final[j][1] = cin_flag
        final[j][2] = score
        
        # save win or lose
        sheet1.cell(j+2, 1).value = j
        sheet1.cell(j+2, 2).value = cin_flag
        sheet1.cell(j+2, 3).value = score
        exl.save('Tuple2.xlsx')

        # show results
        if ((j+1) % milestone) == 0:
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
            print("#Episode: {episode}, score: {score}".format(episode = j+1, score = totalReward // milestone ))
            line = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            filename.write(line)
            line = "#Episode: {episode}, score: {score}\n".format(episode = j+1, score = totalReward // milestone )
            filename.write(line)
            ntuple.saveTuple()
            ntuple.printCount(milestone, filename)
            ntuple.resetCount()
            totalReward = 0
    
    print(winCot / k)
    print(final)
    ntuple.saveTuple()
