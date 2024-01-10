from board import Board
from tuple import Tuple
import numpy as np
import copy
import os
import openpyxl
import csv
import time
import datetime

if __name__ == "__main__":

    currentDatetime = datetime.datetime.now()
    formattedDatetime = currentDatetime.strftime("%m%d-%H%M%S")
    filename = f"Tuple_{formattedDatetime}.xlsx"
    filename1 = f"out_{formattedDatetime}.txt"

    #creat Excel -> record results（win or lose）
    exl = openpyxl.Workbook()
    exl.save(filename)
    exl = openpyxl.load_workbook(filename) 
    sheet1 = exl['Sheet']
    sheet1.cell(1, 1).value = 'Game number' #(row, col) = (1, 1)
    sheet1.cell(1, 2).value = 'Win or lose'
    sheet1.cell(1, 3).value = 'Final Score'
    exl.save(filename)

    # record results（% of tiles）
    filenameOpen = open(filename1, 'w',newline='')
    
    ntuple = Tuple()
    game = Board()
    
    winCot = 0
    loseCot = 0

    list = ["R", "D", "L", "U"]

    k = 3000 #episode
    milestone = 100
    final = np.zeros((k, 3), dtype = np.uint32)
    score200 = []
    score200sum = 0

    totalReward = 0
    
    num_games = 0
    totla_score = 0
    for j in range(k):
        original_tuple = np.copy(ntuple.tuple)
        game.startNewGame()
        score = 0
        cin_flag = 0
        while True:
            tmp_board = game.board
            getT1 = ntuple.getTupleValueSum(game.board)            
            r, game.board = ntuple.takeAction(list, game)
            totalReward += r
            score += r
            #aferstate = game.board()
            game.newTile()
            getT2 = ntuple.getTupleValueSum(game.board)
            ntuple.udpTuple(tmp_board, r+(getT2 - getT1))
            if game.check() == 100:
                cin_flag = 1
            if game.endOfGame() == True:
                break            
        # record result (% of tiles)
        totla_score += score
        num_games += 1     
        if num_games % 10 == 0:
            print("avg_score: ", totla_score/10)
            totla_score = 0
            
        ntuple.addCount(game)  

        # game.printBoard()
        # print("Score: ", score)
        # print("k: ", j)
        
        if cin_flag == 0:
            loseCot += 1
        elif cin_flag == 1:
            winCot += 1

        final[j][0] = j
        final[j][1] = cin_flag
        final[j][2] = score
        
        # save win or lose
        sheet1.cell(j+2, 1).value = j
        sheet1.cell(j+2, 2).value = cin_flag
        sheet1.cell(j+2, 3).value = score
        exl.save('Tuple2.xlsx')

        # show results
        if ((j+1) % milestone) == 0:
            print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
            print("#Episode: {episode}, score: {score}".format(episode = j+1, score = totalReward // milestone ))
            line = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
            filenameOpen.write(line)
            line = "#Episode: {episode}, score: {score}\n".format(episode = j+1, score = totalReward // milestone )
            filenameOpen.write(line)
            ntuple.saveTuple()
            ntuple.printCount(milestone, filenameOpen)
            ntuple.resetCount()
            totalReward = 0
    
    print(winCot / k)
    print(final)
    ntuple.saveTuple()
    exl.close()